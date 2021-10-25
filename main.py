import requests
from bs4 import BeautifulSoup
from datetime import date
from datetime import timedelta
import openpyxl
import linecache as lc
import sys

# get data
day = 0
while 1:
    try:
        t = (date.today() - timedelta(days = day)).strftime("%Y%m%d")
        r = requests.get("https://www.twse.com.tw/exchangeReport/BWIBBU_d?response=html&date=%s&selectType=ALL" % t).text
        soup = BeautifulSoup(r, 'html.parser')
        data = soup.find("div").find("table").find("tbody").get_text().split()
    except:
        day += 1
        continue
    else:
        break

l = len(data) // 7
list = ["證券代號","證券名稱","殖利率(%)","股利年度","本益比","股價淨值比","財報年/季"]
wb = openpyxl.Workbook()
ws = wb.active

# search
if("-s" in sys.argv):
    company = input("search:").split()
    n = 0
    w = -1

    if("-e" in sys.argv):
        # 建立 excel
        table = wb.create_sheet("台股查詢", 0)

        # 寫入 excel
        for i in range(l):
            if data[int(i * 7 + 1)] in company:
                if(w == -1):
                    w = 0
                    for j in range(7):
                        place = chr(65 + w) + str(j+1)
                        table[place] = list[j]
                w += 1
                for j in range(7):
                    place = chr(65 + w) + str(j+1)
                    table[place] = data[n]
                    n += 1
            else:
                n += 7
        if w == -1:
            print("找不到符合條件。")
        else:
            print("save")
            wb.save("查詢結果.xlsx")
    else:
        for i in range(l):
            if data[int(i * 7 + 1)] in company:
                if  w == -1:
                    for i in range(7):
                        print("[%s]" % list[i], end="")
                    print("")
                    w = 0
                for j in range(7):
                    print(data[n], end=" "*(10 - len(str(data[n]))))
                    n += 1
                print("")
            else:
                n += 7
        if w == -1:
            print("找不到符合條件。")
elif("-l" in sys.argv):
    if "-e" in sys.argv:
        table = wb.create_sheet("股票名稱", 0)
        for i in range(l):
            place = chr(65 + i // 48) + str(i+1 - i // 48 * 48)
            table[place] = data[int(i * 7 + 1)]
        print("save")
        wb.save("股票名稱.xlsx")
    else:
        for i in range(l):
            print("%d:" % i,data[int(i * 7 + 1)])
